import openpyxl
import re
import requests
from bs4 import BeautifulSoup

# имя файла для открытия, файл должен быть вместе с программой в корне
book = openpyxl.load_workbook("test.xlsx", )

# выбор активного листа (активный лист является тем, который был открыт на момент сохранения)
sheet = book.active

# ИМЯ ФАЙЛА НА ВЫХОДЕ
OUTPUT_FILE = "output.xlsx"

# regex expression
regex = r"^https:\/\/moscow+\.[A-z0-9]+\.ru+\/catalog+/.*"

# 2 - строка, с которой начинается цикл
# sheet.max_row - предел range
for row in range(2, sheet.max_row):
    table_url = sheet[row][5].value

    # если сайт начинается с https://moscow.petrovich.ru/catalog/ то начинаем работать с парсингом
    match = re.search(regex, str(table_url))
    if table_url != None: # если ссылка существует
        if match: # если ссылка найдена, парсим её
            url = match.group() # сохраняем url 
            response = requests.get(url) # получаем ответ от сервера
            soup = BeautifulSoup(response.text, 'lxml') # генерим соуп
            
            # записываем цену товара
            price_rub = soup.find('p', class_ = 'gold-price')

            # убираем символ валюты
            x = float(''.join(ele for ele in price_rub.text if ele.isdigit() or ele == '.'))
  
            # запись в ячейку
            sheet[f'C{row}'].value = x
            print(f'Стоимость: {x}. Запись в ячейку C{row}')


print(f'Парсинг завершён. Результат сохранён в файле {OUTPUT_FILE}')
book.save(OUTPUT_FILE)